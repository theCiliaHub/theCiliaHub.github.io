#!/usr/bin/env python3
"""
sync_genes.py
=============
Excel → canonical JSON → site JSON files pipeline for CiliaHub gene data.

Source  :  data/source/final list merged human_ciliaminer (3).xlsx
Outputs :  data/generated/genes.json      (canonical dataset)
           ciliahub_data.json             (site search/browse)
           data/genes/ciliAI_master_database.json  (AI chatbot)
           data/genes/ciliAI_lookups.json           (AI lookup index)

Run:
    python3 scripts/sync_genes.py                     # full pipeline
    python3 scripts/sync_genes.py --sheet "5.3.2026"   # explicit sheet
    python3 scripts/sync_genes.py --dry-run             # validate only, no output
    python3 scripts/sync_genes.py --skip-build          # canonical only, skip site JSONs

The script is non-destructive:
  - It writes to a temp file first, then atomically replaces the output.
  - It never modifies the Excel source.
  - Every row-level problem is logged as a WARNING; the row is either fixed or
    skipped, but the script never crashes on bad data.
"""

from __future__ import annotations

import subprocess

import argparse
import hashlib
import json
import logging
import math
import os
import re
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Paths (all relative to the repo root so the script is portable)
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent

# Excel source: prefer the clean name (genes.xlsx), fall back to original.
_EXCEL_CANDIDATES = [
    REPO_ROOT / "data" / "source" / "genes.xlsx",
    REPO_ROOT / "data" / "source" / "final list merged human_ciliaminer (3).xlsx",
]
EXCEL_FILE = next((p for p in _EXCEL_CANDIDATES if p.exists()), _EXCEL_CANDIDATES[0])
OUTPUT_FILE = REPO_ROOT / "data" / "generated" / "genes.json"

# Primary sheet name.  Fall back to the first visible sheet if not found.
DEFAULT_SHEET = "5.3.2026"

# ---------------------------------------------------------------------------
# Column name → canonical field mapping
# ---------------------------------------------------------------------------
# Keys are lowercased, stripped Excel column headers (with common variants).
# Values are the canonical field names used in the output JSON.
COLUMN_MAP: dict[str, str] = {
    "gene":                                                              "name",
    "ensembl_gene_id.x.x":                                              "ensembl_id",
    "ensembl_gene_id":                                                   "ensembl_id",
    "gene.description":                                                  "description",
    "functional.summary.from.literature":                                "functional_summary",
    "synonyms":                                                          "synonym",
    "synonym.":                                                          "synonym",
    "synonym":                                                           "synonym",
    "omim.id":                                                           "omim_id",
    "localization":                                                      "localization",
    "reference":                                                         "reference",
    "protein.complexes":                                                 "protein_complexes",
    "subunits_protein_name":                                             "complex_subunits",
    "protein.complexes referances":                                      "complex_references",
    "gene.annotation":                                                   "gene_annotation",
    "functional.category":                                               "functional_category",
    "pfam_ids":                                                          "pfam_ids",
    "domain_descriptions":                                               "domain_descriptions",
    "ciliopathy":                                                        "ciliopathy",
    "ciliopathy classification":                                         "ciliopathy_classification",
    "disease reference":                                                 "disease_reference",
    "id":                                                                "pathway_id",
    "description":                                                       "pathway_description",
    "source":                                                            "pathway_source",
    "ortholog_mouse":                                                    "ortholog_mouse",
    "ortholog_c_elegans":                                                "ortholog_c_elegans",
    "ortholog_xenopus":                                                  "ortholog_xenopus",
    "ortholog_zebrafish":                                                "ortholog_zebrafish",
    "ortholog_drosophila":                                               "ortholog_drosophila",
    "mouse_ciliopathy_phenotype":                                        "mouse_ciliopathy_phenotype",
    "mouse_phenotype":                                                   "mouse_phenotype",
    "human_ciliopathy_phenotype":                                        "human_ciliopathy_phenotype",
    "human_phenotype":                                                   "human_phenotype",
    "overexpression effects on cilia length (increase/decrease/no effect)":
                                                                         "overexpression_effects",
    "loss-of-function (lof) effects on cilia length (increase/decrease/no effect)":
                                                                         "lof_effects",
    "percentage of ciliated cells (increase/decrease/no effect)":        "percent_ciliated_cells_effects",
}

# Fields that should be split into lists when they contain comma/semicolon separators.
LIST_FIELDS = {
    "localization",
    "functional_category",
    "reference",
    "ciliopathy",
    "pfam_ids",
    "domain_descriptions",
    "ortholog_mouse",
    "ortholog_c_elegans",
    "ortholog_xenopus",
    "ortholog_zebrafish",
    "ortholog_drosophila",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s | %(message)s",
)
log = logging.getLogger("sync_genes")


def _clean_str(value: Any) -> str:
    """Convert a cell value to a clean string, or '' if empty/null/nan."""
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        # Integers stored as floats in Excel (e.g., OMIM IDs)
        if value == math.floor(value):
            return str(int(value))
        return str(value)
    return re.sub(r"[\u200b-\u200d\u2060\ufeff]", "", str(value)).strip()


def _split_list(raw: str) -> list[str]:
    """Split a comma/semicolon-separated string into a deduplicated list."""
    if not raw:
        return []
    parts = re.split(r"[;,]", raw)
    seen: set[str] = set()
    result: list[str] = []
    for p in parts:
        p = p.strip()
        if p and p not in seen:
            seen.add(p)
            result.append(p)
    return result


def _stable_id(gene_name: str, index: int) -> str:
    """
    Return a stable, unique gene identifier.
    Format: GENE<zero-padded-index>  (e.g. GENE00042)
    A content-based hash suffix is appended only for malformed names.
    """
    name = gene_name.strip().upper()
    if re.fullmatch(r"[A-Z0-9\-\.]+", name):
        return f"GENE{index:05d}"
    # Fallback: use first 6 chars of SHA-1 of the raw name
    slug = hashlib.sha1(name.encode()).hexdigest()[:6].upper()
    return f"GENE{index:05d}_{slug}"


def _normalize_header(raw: str | None) -> str:
    """Lower-case, strip, and collapse internal whitespace for a header cell."""
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip().lower())


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def load_excel(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Load an Excel sheet and return (canonical_headers, raw_rows).
    raw_rows is a list of dicts keyed by canonical field names.
    """
    try:
        import openpyxl  # local import so the rest of the script stays importable
    except ImportError:
        log.error("openpyxl is not installed.  Run:  pip install openpyxl")
        sys.exit(1)

    log.info("Opening: %s", path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Sheet selection
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log.info("Using sheet: '%s'", sheet_name)
    else:
        ws = wb.active
        log.warning(
            "Sheet '%s' not found.  Falling back to active sheet: '%s'",
            sheet_name,
            ws.title,
        )

    row_iter = ws.iter_rows(values_only=True)

    # --- Read header row ---
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        log.error("Sheet is empty.")
        sys.exit(1)

    # Build column index → canonical field name mapping
    col_to_field: dict[int, str] = {}
    seen_fields: set[str] = set()
    for col_idx, raw_h in enumerate(raw_headers):
        norm = _normalize_header(raw_h)
        if not norm:
            continue
        field = COLUMN_MAP.get(norm)
        if field is None:
            # Keep unmapped columns under a snake_cased version of the header
            field = re.sub(r"[^a-z0-9]+", "_", norm).strip("_")
            log.debug("Unmapped column '%s' → '%s'", raw_h, field)
        if field in seen_fields:
            # Disambiguate duplicate canonical names
            field = f"{field}_{col_idx}"
            log.debug("Duplicate field resolved to '%s'", field)
        seen_fields.add(field)
        col_to_field[col_idx] = field

    canonical_fields = [col_to_field[i] for i in sorted(col_to_field.keys())]
    log.info("Columns detected: %d mapped, %d total",
             len([f for f in canonical_fields if f in COLUMN_MAP.values()]),
             len(canonical_fields))

    # --- Read data rows ---
    raw_rows: list[dict[str, Any]] = []
    for row in row_iter:
        if all(v is None for v in row):
            continue  # skip blank rows
        record: dict[str, Any] = {}
        for col_idx, field in col_to_field.items():
            cell_val = row[col_idx] if col_idx < len(row) else None
            record[field] = cell_val
        raw_rows.append(record)

    log.info("Raw rows loaded: %d", len(raw_rows))
    return canonical_fields, raw_rows


def validate_and_normalize(
    raw_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    Clean, validate, and normalize each row.
    Returns (clean_genes, stats).
    """
    genes: list[dict[str, Any]] = []
    seen_names: dict[str, int] = {}  # gene_name_upper → first occurrence index
    stats = {
        "total_input": len(raw_rows),
        "skipped_no_name": 0,
        "skipped_duplicate": 0,
        "warnings": 0,
        "output": 0,
    }

    for row_num, raw in enumerate(raw_rows, start=2):  # row 1 = header in Excel
        # ---- 1. Require gene name ----
        name_raw = _clean_str(raw.get("name", ""))
        if not name_raw:
            log.warning("Row %d: missing gene name — skipped.", row_num)
            stats["skipped_no_name"] += 1
            stats["warnings"] += 1
            continue

        # Normalise: strip whitespace, upper-case
        name = name_raw.strip().upper()

        # Basic sanity check: gene symbols should be short alphanumeric tokens
        if len(name) > 50 or not re.search(r"[A-Z0-9]", name):
            log.warning(
                "Row %d: suspicious gene name '%s' — keeping but flagging.",
                row_num, name,
            )
            stats["warnings"] += 1

        # ---- 2. Deduplicate ----
        if name in seen_names:
            log.warning(
                "Row %d: duplicate gene '%s' (first seen at row %d) — skipped.",
                row_num, name, seen_names[name],
            )
            stats["skipped_duplicate"] += 1
            stats["warnings"] += 1
            continue
        seen_names[name] = row_num

        # ---- 3. Build clean record ----
        gene_index = len(genes) + 1
        record: dict[str, Any] = {
            "id":      _stable_id(name, gene_index),
            "name":    name,
            "species": "human",
        }

        # All other string fields
        simple_fields = [
            "ensembl_id", "description", "functional_summary", "synonym",
            "omim_id", "gene_annotation", "protein_complexes", "complex_subunits",
            "complex_references", "pfam_ids", "domain_descriptions",
            "ciliopathy_classification", "disease_reference",
            "pathway_id", "pathway_description", "pathway_source",
            "ortholog_mouse", "ortholog_c_elegans", "ortholog_xenopus",
            "ortholog_zebrafish", "ortholog_drosophila",
            "mouse_ciliopathy_phenotype", "mouse_phenotype",
            "human_ciliopathy_phenotype", "human_phenotype",
            "overexpression_effects", "lof_effects",
            "percent_ciliated_cells_effects",
        ]

        for field in simple_fields:
            raw_val = _clean_str(raw.get(field, ""))
            if field in LIST_FIELDS:
                record[field] = _split_list(raw_val)
            else:
                record[field] = raw_val if raw_val else None

        # List fields already handled above; add remaining ones explicitly
        for list_field in LIST_FIELDS:
            if list_field not in simple_fields and list_field in raw:
                raw_val = _clean_str(raw.get(list_field, ""))
                record[list_field] = _split_list(raw_val)

        # localization and functional_category get list treatment
        for field in ("localization", "functional_category", "reference", "ciliopathy"):
            raw_val = _clean_str(raw.get(field, ""))
            record[field] = _split_list(raw_val)

        genes.append(record)

    stats["output"] = len(genes)
    return genes, stats


def write_output(genes: list[dict[str, Any]], path: Path, dry_run: bool) -> None:
    """Atomically write the canonical JSON dataset."""
    payload = {
        "meta": {
            "generated_at":     datetime.now(timezone.utc).isoformat(),
            "generator":        "scripts/sync_genes.py",
            "source":           str(EXCEL_FILE.relative_to(REPO_ROOT)),
            "total_genes":      len(genes),
            "schema_version":   "1.0",
        },
        "genes": genes,
    }

    if dry_run:
        log.info("[DRY-RUN] Would write %d genes to %s", len(genes), path)
        # Still validate JSON serialisability
        json.dumps(payload, ensure_ascii=False)
        log.info("[DRY-RUN] JSON serialisation OK.")
        return

    path.parent.mkdir(parents=True, exist_ok=True)

    # Write to a temp file in the same directory, then rename (atomic on POSIX)
    fd, tmp_path = tempfile.mkstemp(dir=path.parent, suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write("\n")  # POSIX trailing newline
        os.replace(tmp_path, path)  # atomic rename
        log.info("✅  Written: %s  (%d genes)", path, len(genes))
    except Exception:
        # Clean up orphaned temp file on error
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync CiliaHub Excel gene list → canonical JSON dataset.",
    )
    parser.add_argument(
        "--excel",
        default=str(EXCEL_FILE),
        help="Path to the source Excel file (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        default=str(OUTPUT_FILE),
        help="Path for the generated JSON output (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help="Excel sheet name to read (default: %(default)s)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and report only — do not write any output file.",
    )
    parser.add_argument(
        "--skip-build",
        action="store_true",
        help="Stop after writing the canonical JSON; do not regenerate site JSON files.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable DEBUG-level logging.",
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    excel_path = Path(args.excel)
    output_path = Path(args.output)

    if not excel_path.exists():
        log.error("Excel file not found: %s", excel_path)
        sys.exit(1)

    # --- Pipeline ---
    _fields, raw_rows = load_excel(excel_path, args.sheet)
    genes, stats = validate_and_normalize(raw_rows)

    # --- Summary report ---
    log.info("─" * 50)
    log.info("Input rows   : %d", stats["total_input"])
    log.info("Skipped (no name)   : %d", stats["skipped_no_name"])
    log.info("Skipped (duplicate) : %d", stats["skipped_duplicate"])
    log.info("Total warnings      : %d", stats["warnings"])
    log.info("Output genes        : %d", stats["output"])
    log.info("─" * 50)

    write_output(genes, output_path, dry_run=args.dry_run)

    if stats["warnings"] > 0:
        log.warning("%d warning(s) encountered — review the log above.", stats["warnings"])
    else:
        log.info("No warnings.")

    # ── Auto-invoke build_site_jsons.py unless skipped ────────────────────────
    if not args.dry_run and not args.skip_build:
        build_script = Path(__file__).parent / "build_site_jsons.py"
        if build_script.exists():
            log.info("")
            log.info("─" * 50)
            log.info("Running build_site_jsons.py to regenerate site files...")
            log.info("─" * 50)
            cmd = [sys.executable, str(build_script)]
            if args.verbose:
                cmd.append("--verbose")
            result = subprocess.run(cmd, check=False)
            if result.returncode != 0:
                log.error("build_site_jsons.py exited with code %d", result.returncode)
                sys.exit(result.returncode)
        else:
            log.warning("build_site_jsons.py not found at %s — skipping site JSON rebuild.", build_script)


if __name__ == "__main__":
    main()
